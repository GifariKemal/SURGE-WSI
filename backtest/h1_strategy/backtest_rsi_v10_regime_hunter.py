"""
RSI v3.7 V10 - "The Regime Hunter" Backtest
============================================
Testing Gemini's recommendations to beat V5:

1. ADX Filter (The Shield): Skip entry if ADX > 30 (strong trend)
2. Confirmation Candle: Entry only after reversal candle
3. ATR-Based Dynamic Lot Size: Scale lot inversely with volatility
4. Trailing Stop: Lock profits in winning trades

Control Group: V5 baseline
Test Groups: V10 variants

Author: SURGE-WSI Team
"""
import os
import MetaTrader5 as mt5
import pandas as pd
import numpy as np
from datetime import datetime, timezone
from dataclasses import dataclass
from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

MT5_LOGIN = int(os.getenv('MT5_LOGIN', '61045904'))
MT5_PASSWORD = os.getenv('MT5_PASSWORD', '')
MT5_SERVER = os.getenv('MT5_SERVER', 'FinexBisnisSolusi-Demo')
MT5_PATH = os.getenv('MT5_PATH', r"C:\Program Files\Finex Bisnis Solusi MT5 Terminal\terminal64.exe")

SYMBOL = "GBPUSD"
INITIAL_BALANCE = 10000.0
PIP_SIZE = 0.0001
PIP_VALUE = 10.0  # per standard lot


# =============================================================================
# STRATEGY CONFIGURATIONS
# =============================================================================

@dataclass
class StrategyConfig:
    """Strategy configuration"""
    name: str

    # RSI Settings
    rsi_period: int = 10
    rsi_oversold: int = 42
    rsi_overbought: int = 58

    # ATR Settings
    atr_period: int = 14
    min_atr_pct: int = 20
    max_atr_pct: int = 80

    # TP/SL Settings
    sl_mult: float = 1.5
    tp_low: float = 2.4
    tp_med: float = 3.0
    tp_high: float = 3.6

    # Risk Management
    risk_per_trade: float = 0.01
    max_holding_hours: int = 46

    # Regime Filter (V5 style)
    use_regime_filter: bool = True

    # Consecutive Loss Filter
    use_consec_loss_filter: bool = True
    consec_loss_limit: int = 3

    # === V10 NEW FEATURES ===

    # ADX Filter (The Shield)
    use_adx_filter: bool = False
    adx_period: int = 14
    max_adx_for_entry: int = 30

    # Confirmation Candle
    use_candle_confirm: bool = False

    # Dynamic ATR-Based Lot Size
    use_dynamic_lot: bool = False

    # Trailing Stop
    use_trailing_stop: bool = False
    trailing_start_atr: float = 1.5
    trailing_step_atr: float = 0.5


# Pre-defined configurations
CONFIGS = {
    'V5_Baseline': StrategyConfig(
        name='V5_Baseline',
        rsi_oversold=42,
        rsi_overbought=58,
        use_regime_filter=True,
        use_consec_loss_filter=True,
    ),
    'V10A_ADX_Only': StrategyConfig(
        name='V10A_ADX_Only',
        rsi_oversold=42,
        rsi_overbought=58,
        use_regime_filter=True,
        use_consec_loss_filter=True,
        use_adx_filter=True,
        max_adx_for_entry=30,
    ),
    'V10B_ADX_Confirm': StrategyConfig(
        name='V10B_ADX_Confirm',
        rsi_oversold=42,
        rsi_overbought=58,
        use_regime_filter=True,
        use_consec_loss_filter=True,
        use_adx_filter=True,
        max_adx_for_entry=30,
        use_candle_confirm=True,
    ),
    'V10C_ADX_Confirm_DynLot': StrategyConfig(
        name='V10C_ADX_Confirm_DynLot',
        rsi_oversold=42,
        rsi_overbought=58,
        use_regime_filter=True,
        use_consec_loss_filter=True,
        use_adx_filter=True,
        max_adx_for_entry=30,
        use_candle_confirm=True,
        use_dynamic_lot=True,
    ),
    'V10D_Full': StrategyConfig(
        name='V10D_Full',
        rsi_oversold=40,  # Tightened per Gemini
        rsi_overbought=60,  # Tightened per Gemini
        use_regime_filter=True,
        use_consec_loss_filter=True,
        use_adx_filter=True,
        max_adx_for_entry=30,
        use_candle_confirm=True,
        use_dynamic_lot=True,
        use_trailing_stop=True,
        trailing_start_atr=1.5,
        trailing_step_atr=0.5,
        risk_per_trade=0.015,  # Increased due to tighter filters
    ),
    'V10E_ADX25': StrategyConfig(
        name='V10E_ADX25',
        rsi_oversold=40,
        rsi_overbought=60,
        use_regime_filter=True,
        use_consec_loss_filter=True,
        use_adx_filter=True,
        max_adx_for_entry=25,  # More conservative
        use_candle_confirm=True,
        use_dynamic_lot=True,
        use_trailing_stop=True,
    ),
}


# =============================================================================
# MT5 FUNCTIONS
# =============================================================================

def connect_mt5():
    if not MT5_PASSWORD:
        print("ERROR: MT5_PASSWORD not set")
        return False
    if not mt5.initialize(path=MT5_PATH):
        print(f"MT5 init failed: {mt5.last_error()}")
        return False
    if not mt5.login(MT5_LOGIN, MT5_PASSWORD, MT5_SERVER):
        print(f"MT5 login failed: {mt5.last_error()}")
        return False
    acc = mt5.account_info()
    print(f"Connected: {acc.login} | Balance: ${acc.balance:,.2f}")
    return True


def get_h1_data(symbol, start_date, end_date):
    rates = mt5.copy_rates_range(symbol, mt5.TIMEFRAME_H1, start_date, end_date)
    if rates is None:
        return None
    df = pd.DataFrame(rates)
    df['time'] = pd.to_datetime(df['time'], unit='s')
    df.set_index('time', inplace=True)
    return df


# =============================================================================
# INDICATORS
# =============================================================================

def calculate_rsi(series: pd.Series, period: int) -> pd.Series:
    """Calculate RSI"""
    delta = series.diff()
    gain = delta.where(delta > 0, 0).rolling(period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(period).mean()
    rs = np.where(loss == 0, 100, gain / loss)
    return pd.Series(100 - (100 / (1 + rs)), index=series.index)


def calculate_atr(df: pd.DataFrame, period: int) -> pd.Series:
    """Calculate ATR"""
    tr = np.maximum(
        df['high'] - df['low'],
        np.maximum(
            abs(df['high'] - df['close'].shift(1)),
            abs(df['low'] - df['close'].shift(1))
        )
    )
    return pd.Series(tr, index=df.index).rolling(period).mean()


def calculate_adx(df: pd.DataFrame, period: int = 14) -> pd.Series:
    """Calculate ADX indicator"""
    plus_dm = (df['high'] - df['high'].shift(1)).clip(lower=0)
    minus_dm = (df['low'].shift(1) - df['low']).clip(lower=0)

    # When +DM > -DM, -DM = 0 and vice versa
    plus_dm = np.where(plus_dm > minus_dm, plus_dm, 0)
    minus_dm = np.where(minus_dm > plus_dm, minus_dm, 0)

    tr = np.maximum(
        df['high'] - df['low'],
        np.maximum(
            abs(df['high'] - df['close'].shift(1)),
            abs(df['low'] - df['close'].shift(1))
        )
    )

    atr = pd.Series(tr, index=df.index).rolling(period).mean()
    plus_di = 100 * pd.Series(plus_dm, index=df.index).rolling(period).mean() / atr
    minus_di = 100 * pd.Series(minus_dm, index=df.index).rolling(period).mean() / atr

    dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di + 0.0001)
    adx = dx.rolling(period).mean()

    return adx


def calculate_atr_percentile(atr_series: pd.Series, lookback: int = 100) -> pd.Series:
    """Calculate ATR percentile"""
    def pct_func(x):
        if len(x) < 2:
            return 50.0
        current = x[-1]
        count_below = (x[:-1] < current).sum()
        return (count_below / (len(x) - 1)) * 100

    return atr_series.rolling(lookback).apply(pct_func, raw=True)


def prepare_indicators(df: pd.DataFrame, config: StrategyConfig) -> pd.DataFrame:
    """Calculate all indicators based on config"""

    # RSI
    df['rsi'] = calculate_rsi(df['close'], config.rsi_period)

    # ATR
    df['atr'] = calculate_atr(df, config.atr_period)
    df['atr_pct'] = calculate_atr_percentile(df['atr'])

    # ADX (for V10)
    df['adx'] = calculate_adx(df, config.adx_period)

    # SMAs for regime detection
    df['sma_20'] = df['close'].rolling(20).mean()
    df['sma_50'] = df['close'].rolling(50).mean()
    df['sma_slope'] = (df['sma_20'] / df['sma_20'].shift(10) - 1) * 100

    # Regime Detection
    conditions = [
        (df['sma_20'] > df['sma_50']) & (df['sma_slope'] > 0.5),
        (df['sma_20'] < df['sma_50']) & (df['sma_slope'] < -0.5),
    ]
    df['regime'] = np.select(conditions, ['BULL', 'BEAR'], default='SIDEWAYS')

    # Candle patterns for confirmation
    df['is_bullish'] = df['close'] > df['open']
    df['is_bearish'] = df['close'] < df['open']
    df['body_size'] = abs(df['close'] - df['open'])
    df['prev_close'] = df['close'].shift(1)

    # Time features
    df['hour'] = df.index.hour
    df['weekday'] = df.index.weekday

    return df.ffill().fillna(0)


# =============================================================================
# BACKTEST ENGINE
# =============================================================================

@dataclass
class Trade:
    ticket: int
    entry_time: datetime
    exit_time: Optional[datetime]
    direction: int  # 1 = BUY, -1 = SELL
    entry_price: float
    sl_price: float
    tp_price: float
    lot_size: float
    exit_price: Optional[float] = None
    pnl: float = 0.0
    pips: float = 0.0
    exit_reason: str = ""
    rsi: float = 0.0
    atr_pct: float = 0.0
    adx: float = 0.0


def check_confirmation_candle(df: pd.DataFrame, idx: int, signal: int) -> bool:
    """
    Check if current candle confirms the signal direction
    BUY: Current close > Previous close (bullish reversal)
    SELL: Current close < Previous close (bearish reversal)
    """
    if idx < 1:
        return False

    current = df.iloc[idx]
    prev = df.iloc[idx - 1]

    if signal == 1:  # BUY
        # Bullish confirmation: close > prev close AND bullish candle
        return current['close'] > prev['close'] and current['is_bullish']
    else:  # SELL
        # Bearish confirmation: close < prev close AND bearish candle
        return current['close'] < prev['close'] and current['is_bearish']


def calculate_dynamic_lot_size(balance: float, atr: float, sl_mult: float,
                                risk_pct: float, base_atr: float) -> float:
    """
    Calculate lot size inversely proportional to ATR
    Higher ATR = smaller lot, Lower ATR = larger lot

    Formula: Lot = Risk$ / (ATR * SL_Mult * PIP_VALUE)
    Adjusted by ATR ratio to baseline
    """
    risk_amount = balance * risk_pct
    sl_pips = atr / PIP_SIZE * sl_mult

    if sl_pips <= 0:
        return 0.01

    # Base lot size from risk
    base_lot = risk_amount / (sl_pips * PIP_VALUE)

    # ATR adjustment factor (inverse relationship)
    # If current ATR > base_atr, reduce lot; if < base_atr, increase lot
    if base_atr > 0:
        atr_ratio = base_atr / atr  # Inverse: high ATR = small ratio
        atr_ratio = max(0.5, min(1.5, atr_ratio))  # Clamp to 0.5x - 1.5x
        adjusted_lot = base_lot * atr_ratio
    else:
        adjusted_lot = base_lot

    return max(0.01, min(5.0, round(adjusted_lot, 2)))


def run_backtest(df: pd.DataFrame, config: StrategyConfig,
                 test_start: str = '2024-10-01', test_end: str = '2026-02-01') -> Dict:
    """Run backtest with specified configuration"""

    balance = INITIAL_BALANCE
    wins = losses = 0
    position: Optional[Trade] = None
    peak = balance
    max_dd = 0
    max_dd_amount = 0
    trades: List[Trade] = []
    consecutive_losses = 0
    max_consec_losses = 0
    trades_filtered = 0
    filter_reasons = {
        'regime': 0,
        'consec_loss': 0,
        'adx': 0,
        'candle_confirm': 0,
        'atr_range': 0,
        'hours': 0,
    }

    # Calculate baseline ATR for dynamic lot sizing
    base_atr = df['atr'].iloc[200:500].mean()

    monthly_pnl = {}

    for i in range(200, len(df) - 20):
        row = df.iloc[i]
        current_time = df.index[i]
        in_test = current_time >= pd.Timestamp(test_start) and current_time < pd.Timestamp(test_end)

        if not in_test:
            continue

        month_str = current_time.strftime('%Y-%m')
        weekday = row['weekday']
        hour = row['hour']

        if weekday >= 5:
            continue

        # === Position Management ===
        if position:
            exit_reason = None
            exit_price = None
            pnl = 0

            # Timeout check
            hours_held = (i - position.entry_time.hour) if hasattr(position.entry_time, 'hour') else 0
            # Approximate hours held
            entry_idx = next((idx for idx, t in enumerate(df.index) if t == position.entry_time), 0)
            hours_held = i - entry_idx

            if hours_held >= config.max_holding_hours:
                exit_price = row['close']
                exit_reason = 'TIMEOUT'
            else:
                # Check SL/TP
                if position.direction == 1:  # BUY
                    if row['low'] <= position.sl_price:
                        exit_price = position.sl_price
                        exit_reason = 'SL'
                    elif row['high'] >= position.tp_price:
                        exit_price = position.tp_price
                        exit_reason = 'TP'
                    # Trailing stop logic (V10)
                    elif config.use_trailing_stop:
                        current_profit_atr = (row['high'] - position.entry_price) / row['atr']
                        if current_profit_atr >= config.trailing_start_atr:
                            new_sl = row['high'] - (row['atr'] * config.trailing_step_atr)
                            if new_sl > position.sl_price:
                                position.sl_price = new_sl
                else:  # SELL
                    if row['high'] >= position.sl_price:
                        exit_price = position.sl_price
                        exit_reason = 'SL'
                    elif row['low'] <= position.tp_price:
                        exit_price = position.tp_price
                        exit_reason = 'TP'
                    # Trailing stop logic (V10)
                    elif config.use_trailing_stop:
                        current_profit_atr = (position.entry_price - row['low']) / row['atr']
                        if current_profit_atr >= config.trailing_start_atr:
                            new_sl = row['low'] + (row['atr'] * config.trailing_step_atr)
                            if new_sl < position.sl_price:
                                position.sl_price = new_sl

            if exit_reason:
                if position.direction == 1:
                    pips = (exit_price - position.entry_price) / PIP_SIZE
                else:
                    pips = (position.entry_price - exit_price) / PIP_SIZE

                pnl = pips * position.lot_size * PIP_VALUE
                balance += pnl

                if pnl > 0:
                    wins += 1
                    consecutive_losses = 0
                else:
                    losses += 1
                    consecutive_losses += 1
                    if consecutive_losses > max_consec_losses:
                        max_consec_losses = consecutive_losses

                position.exit_time = current_time
                position.exit_price = exit_price
                position.pnl = pnl
                position.pips = pips
                position.exit_reason = exit_reason
                trades.append(position)

                if month_str not in monthly_pnl:
                    monthly_pnl[month_str] = {'pnl': 0, 'trades': 0, 'wins': 0}
                monthly_pnl[month_str]['pnl'] += pnl
                monthly_pnl[month_str]['trades'] += 1
                if pnl > 0:
                    monthly_pnl[month_str]['wins'] += 1

                position = None

            # Update drawdown
            if balance > peak:
                peak = balance
            dd = (peak - balance) / peak * 100
            if dd > max_dd:
                max_dd = dd
                max_dd_amount = peak - balance

            continue  # Skip entry if we just managed position

        # === Entry Logic with Filters ===

        # Trading hours filter
        if hour < 7 or hour >= 22 or hour == 12:
            filter_reasons['hours'] += 1
            continue

        # ATR percentile filter
        atr_pct = row['atr_pct']
        if atr_pct < config.min_atr_pct or atr_pct > config.max_atr_pct:
            filter_reasons['atr_range'] += 1
            continue

        # === V5 Filters ===

        # Regime filter (SIDEWAYS only)
        if config.use_regime_filter:
            if row['regime'] != 'SIDEWAYS':
                filter_reasons['regime'] += 1
                trades_filtered += 1
                continue

        # Consecutive loss filter
        if config.use_consec_loss_filter:
            if consecutive_losses >= config.consec_loss_limit:
                filter_reasons['consec_loss'] += 1
                trades_filtered += 1
                consecutive_losses = 0  # Reset after pause
                continue

        # === V10 NEW FILTERS ===

        # ADX Filter (The Shield)
        if config.use_adx_filter:
            if row['adx'] > config.max_adx_for_entry:
                filter_reasons['adx'] += 1
                trades_filtered += 1
                continue

        # RSI Signal
        rsi = row['rsi']
        signal = 0
        if rsi < config.rsi_oversold:
            signal = 1  # BUY
        elif rsi > config.rsi_overbought:
            signal = -1  # SELL

        if signal == 0:
            continue

        # Confirmation Candle Filter (V10)
        if config.use_candle_confirm:
            if not check_confirmation_candle(df, i, signal):
                filter_reasons['candle_confirm'] += 1
                trades_filtered += 1
                continue

        # === Execute Trade ===
        entry_price = row['close']
        atr = row['atr'] if row['atr'] > 0 else entry_price * 0.002

        # Dynamic TP based on ATR percentile
        if atr_pct < 40:
            tp_mult = config.tp_low
        elif atr_pct > 60:
            tp_mult = config.tp_high
        else:
            tp_mult = config.tp_med

        # Session bonus (12-16 overlap)
        if 12 <= hour < 16:
            tp_mult += 0.35

        # Calculate SL/TP levels
        if signal == 1:  # BUY
            sl_price = entry_price - (atr * config.sl_mult)
            tp_price = entry_price + (atr * tp_mult)
        else:  # SELL
            sl_price = entry_price + (atr * config.sl_mult)
            tp_price = entry_price - (atr * tp_mult)

        # Position sizing
        if config.use_dynamic_lot:
            lot_size = calculate_dynamic_lot_size(
                balance, atr, config.sl_mult, config.risk_per_trade, base_atr
            )
        else:
            risk_amount = balance * config.risk_per_trade
            sl_pips = abs(entry_price - sl_price) / PIP_SIZE
            lot_size = risk_amount / (sl_pips * PIP_VALUE) if sl_pips > 0 else 0.01
            lot_size = max(0.01, min(5.0, round(lot_size, 2)))

        position = Trade(
            ticket=len(trades) + 1,
            entry_time=current_time,
            exit_time=None,
            direction=signal,
            entry_price=entry_price,
            sl_price=sl_price,
            tp_price=tp_price,
            lot_size=lot_size,
            rsi=rsi,
            atr_pct=atr_pct,
            adx=row['adx']
        )

    # Calculate final stats
    total_trades = wins + losses
    win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
    total_return = ((balance - INITIAL_BALANCE) / INITIAL_BALANCE) * 100

    winning_trades = [t for t in trades if t.pnl > 0]
    losing_trades = [t for t in trades if t.pnl <= 0]

    gross_profit = sum(t.pnl for t in winning_trades)
    gross_loss = abs(sum(t.pnl for t in losing_trades))
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float('inf')

    avg_win = gross_profit / len(winning_trades) if winning_trades else 0
    avg_loss = gross_loss / len(losing_trades) if losing_trades else 0

    # Calculate Sharpe Ratio (approximate)
    if trades:
        returns = [t.pnl for t in trades]
        avg_return = np.mean(returns)
        std_return = np.std(returns) if len(returns) > 1 else 1
        sharpe = (avg_return / std_return) * np.sqrt(252) if std_return > 0 else 0
    else:
        sharpe = 0

    return {
        'config_name': config.name,
        'balance': balance,
        'total_return': total_return,
        'net_profit': balance - INITIAL_BALANCE,
        'max_drawdown': max_dd,
        'max_dd_amount': max_dd_amount,
        'total_trades': total_trades,
        'wins': wins,
        'losses': losses,
        'win_rate': win_rate,
        'profit_factor': profit_factor,
        'sharpe_ratio': sharpe,
        'gross_profit': gross_profit,
        'gross_loss': gross_loss,
        'avg_win': avg_win,
        'avg_loss': avg_loss,
        'max_consec_losses': max_consec_losses,
        'trades_filtered': trades_filtered,
        'filter_reasons': filter_reasons,
        'monthly_pnl': monthly_pnl,
        'trades': trades,
    }


def print_comparison(results: List[Dict]):
    """Print comparison table of all strategies"""

    print("\n" + "=" * 100)
    print("STRATEGY COMPARISON - V10 'The Regime Hunter' vs V5 Baseline")
    print("=" * 100)

    # Header
    print(f"\n{'Strategy':<25} {'Net Profit':>12} {'Return':>10} {'DD%':>8} {'Trades':>8} {'WR%':>8} {'PF':>8} {'Sharpe':>8} {'MaxCL':>8}")
    print("-" * 100)

    # Sort by net profit
    sorted_results = sorted(results, key=lambda x: x['net_profit'], reverse=True)

    for r in sorted_results:
        print(f"{r['config_name']:<25} ${r['net_profit']:>10,.0f} {r['total_return']:>9.1f}% {r['max_drawdown']:>7.1f}% {r['total_trades']:>8} {r['win_rate']:>7.1f}% {r['profit_factor']:>7.2f} {r['sharpe_ratio']:>7.2f} {r['max_consec_losses']:>8}")

    print("-" * 100)

    # Find best strategy
    best = sorted_results[0]
    v5 = next((r for r in results if r['config_name'] == 'V5_Baseline'), None)

    print(f"\nBEST STRATEGY: {best['config_name']}")
    print(f"  Net Profit: ${best['net_profit']:,.2f}")
    print(f"  Return: {best['total_return']:.1f}%")
    print(f"  Max Drawdown: {best['max_drawdown']:.1f}%")
    print(f"  Profit Factor: {best['profit_factor']:.2f}")

    if v5 and best['config_name'] != 'V5_Baseline':
        improvement = best['net_profit'] - v5['net_profit']
        dd_improvement = v5['max_drawdown'] - best['max_drawdown']
        print(f"\n  vs V5 Baseline:")
        print(f"    Profit Improvement: ${improvement:+,.2f}")
        print(f"    Drawdown Improvement: {dd_improvement:+.1f}%")

    # Print filter effectiveness for V10 variants
    print("\n" + "=" * 100)
    print("FILTER EFFECTIVENESS")
    print("=" * 100)

    for r in sorted_results:
        if 'V10' in r['config_name']:
            print(f"\n{r['config_name']}:")
            print(f"  Total Filtered: {r['trades_filtered']}")
            for reason, count in r['filter_reasons'].items():
                if count > 0:
                    print(f"    - {reason}: {count}")


def export_comparison_excel(results: List[Dict], filename: str):
    """Export comparison results to Excel"""

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    profit_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    loss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    best_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

    # Sheet 1: Summary Comparison
    ws = wb.active
    ws.title = "Comparison"

    ws['A1'] = "V10 'THE REGIME HUNTER' - STRATEGY COMPARISON"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')

    headers = ['Strategy', 'Net Profit', 'Return %', 'Max DD %', 'Trades', 'Win Rate %',
               'Profit Factor', 'Sharpe', 'Max Consec Loss', 'Filtered']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    sorted_results = sorted(results, key=lambda x: x['net_profit'], reverse=True)

    for row_idx, r in enumerate(sorted_results, 4):
        ws.cell(row=row_idx, column=1, value=r['config_name'])
        ws.cell(row=row_idx, column=2, value=r['net_profit'])
        ws.cell(row=row_idx, column=3, value=r['total_return'])
        ws.cell(row=row_idx, column=4, value=r['max_drawdown'])
        ws.cell(row=row_idx, column=5, value=r['total_trades'])
        ws.cell(row=row_idx, column=6, value=r['win_rate'])
        ws.cell(row=row_idx, column=7, value=r['profit_factor'])
        ws.cell(row=row_idx, column=8, value=r['sharpe_ratio'])
        ws.cell(row=row_idx, column=9, value=r['max_consec_losses'])
        ws.cell(row=row_idx, column=10, value=r['trades_filtered'])

        # Highlight best
        if row_idx == 4:
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = best_fill

    # Adjust column widths
    widths = [25, 12, 10, 10, 8, 12, 12, 10, 15, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Sheet 2: Monthly Comparison
    ws2 = wb.create_sheet("Monthly")

    ws2['A1'] = "MONTHLY P/L COMPARISON"
    ws2['A1'].font = Font(bold=True, size=14)

    # Get all months
    all_months = set()
    for r in results:
        all_months.update(r['monthly_pnl'].keys())
    all_months = sorted(all_months)

    # Headers
    ws2.cell(row=3, column=1, value="Month").font = header_font
    ws2.cell(row=3, column=1).fill = header_fill

    for col, r in enumerate(sorted_results, 2):
        ws2.cell(row=3, column=col, value=r['config_name']).font = header_font
        ws2.cell(row=3, column=col).fill = header_fill

    # Data
    for row_idx, month in enumerate(all_months, 4):
        ws2.cell(row=row_idx, column=1, value=month)

        for col, r in enumerate(sorted_results, 2):
            pnl = r['monthly_pnl'].get(month, {}).get('pnl', 0)
            cell = ws2.cell(row=row_idx, column=col, value=pnl)
            cell.number_format = '#,##0.00'

            if pnl >= 0:
                cell.fill = profit_fill
            else:
                cell.fill = loss_fill

    wb.save(filename)
    print(f"\nComparison report saved: {filename}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("RSI v3.7 V10 - 'THE REGIME HUNTER' Backtest")
    print("Testing ADX Filter + Confirmation Candle + Dynamic Lot")
    print("=" * 70)

    if not connect_mt5():
        return

    try:
        print(f"\nFetching {SYMBOL} H1 data...")
        start_date = datetime(2024, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(2026, 1, 31, 23, 59, tzinfo=timezone.utc)
        df = get_h1_data(SYMBOL, start_date, end_date)

        if df is None or len(df) == 0:
            print("Failed to get data")
            return

        print(f"Loaded {len(df)} bars")

        results = []

        for name, config in CONFIGS.items():
            print(f"\nPreparing indicators for {name}...")
            df_copy = df.copy()
            df_copy = prepare_indicators(df_copy, config)

            print(f"Running backtest: {name}...")
            result = run_backtest(df_copy, config)
            results.append(result)

            print(f"  -> Net Profit: ${result['net_profit']:,.0f} | "
                  f"Return: {result['total_return']:.1f}% | "
                  f"DD: {result['max_drawdown']:.1f}% | "
                  f"Trades: {result['total_trades']} | "
                  f"PF: {result['profit_factor']:.2f}")

        # Print comparison
        print_comparison(results)

        # Export to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"C:/Users/Administrator/Music/SURGE-WSI/backtest/h1_strategy/V10_Comparison_Report_{timestamp}.xlsx"
        export_comparison_excel(results, filename)

        print("\n" + "=" * 70)
        print("BACKTEST COMPLETE")
        print("=" * 70)

    finally:
        mt5.shutdown()


if __name__ == "__main__":
    main()
