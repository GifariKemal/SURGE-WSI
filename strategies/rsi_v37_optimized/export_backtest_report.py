"""
RSI v3.7 OPTIMIZED - Export Backtest Report to Excel
=====================================================
Exports backtest results in MT5 Strategy Tester format
"""
import os
import MetaTrader5 as mt5
import pandas as pd
import numpy as np
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

MT5_LOGIN = int(os.getenv('MT5_LOGIN', '61045904'))
MT5_PASSWORD = os.getenv('MT5_PASSWORD', '')
MT5_SERVER = os.getenv('MT5_SERVER', 'FinexBisnisSolusi-Demo')
MT5_PATH = os.getenv('MT5_PATH', r"C:\Program Files\Finex Bisnis Solusi MT5 Terminal\terminal64.exe")

SYMBOL = "GBPUSD"
INITIAL_BALANCE = 10000.0

# Strategy Parameters
RSI_PERIOD = 10
RSI_OVERSOLD = 42
RSI_OVERBOUGHT = 58
ATR_PERIOD = 14
SL_MULT = 1.5
TP_LOW = 2.4
TP_MED = 3.0
TP_HIGH = 3.6
MAX_HOLDING_HOURS = 46
MIN_ATR_PCT = 20
MAX_ATR_PCT = 80
RISK_PER_TRADE = 0.01

# Filter Settings
USE_REGIME_FILTER = True
ALLOWED_REGIMES = ['SIDEWAYS']
USE_CONSEC_LOSS_FILTER = True
CONSEC_LOSS_LIMIT = 3


# =============================================================================
# MT5 FUNCTIONS
# =============================================================================

def connect_mt5():
    if not MT5_PASSWORD:
        print("ERROR: MT5_PASSWORD not set")
        return False
    if not mt5.initialize(path=MT5_PATH):
        print(f"MT5 init failed: {mt5.last_error()}")
        return False
    if not mt5.login(MT5_LOGIN, MT5_PASSWORD, MT5_SERVER):
        print(f"MT5 login failed: {mt5.last_error()}")
        return False
    acc = mt5.account_info()
    print(f"Connected: {acc.login} | Balance: ${acc.balance:,.2f}")
    return True


def get_h1_data(symbol, start_date, end_date):
    rates = mt5.copy_rates_range(symbol, mt5.TIMEFRAME_H1, start_date, end_date)
    if rates is None:
        return None
    df = pd.DataFrame(rates)
    df['time'] = pd.to_datetime(df['time'], unit='s')
    df.set_index('time', inplace=True)
    return df


# =============================================================================
# INDICATORS
# =============================================================================

def prepare_indicators(df):
    """Calculate all indicators."""

    # RSI
    delta = df['close'].diff()
    gain = delta.where(delta > 0, 0).rolling(RSI_PERIOD).mean()
    loss_s = (-delta.where(delta < 0, 0)).rolling(RSI_PERIOD).mean()
    rs = np.where(loss_s == 0, 100, gain / loss_s)
    df['rsi'] = 100 - (100 / (1 + rs))

    # ATR
    tr = np.maximum(df['high'] - df['low'],
                    np.maximum(abs(df['high'] - df['close'].shift(1)),
                              abs(df['low'] - df['close'].shift(1))))
    df['atr'] = pd.Series(tr, index=df.index).rolling(ATR_PERIOD).mean()

    # ATR Percentile
    def atr_pct_func(x):
        if len(x) == 0:
            return 50.0
        current = x[-1]
        count_below = (x[:-1] < current).sum()
        return (count_below / (len(x) - 1)) * 100 if len(x) > 1 else 50.0
    df['atr_pct'] = df['atr'].rolling(100).apply(atr_pct_func, raw=True)

    # SMAs for regime
    df['sma_20'] = df['close'].rolling(20).mean()
    df['sma_50'] = df['close'].rolling(50).mean()
    df['sma_slope'] = (df['sma_20'] / df['sma_20'].shift(10) - 1) * 100

    # Regime Detection
    conditions = [
        (df['sma_20'] > df['sma_50']) & (df['sma_slope'] > 0.5),
        (df['sma_20'] < df['sma_50']) & (df['sma_slope'] < -0.5),
    ]
    df['regime'] = np.select(conditions, ['BULL', 'BEAR'], default='SIDEWAYS')

    # Time features
    df['hour'] = df.index.hour
    df['weekday'] = df.index.weekday

    return df.ffill().fillna(0)


# =============================================================================
# BACKTEST ENGINE
# =============================================================================

def run_backtest(df, test_start='2024-10-01', test_end='2026-02-01'):
    """Run RSI v3.7 OPTIMIZED backtest with detailed trade logging."""

    balance = INITIAL_BALANCE
    wins = losses = 0
    position = None
    peak = balance
    max_dd = 0
    max_dd_amount = 0
    monthly_pnl = {}
    trades_taken = 0
    trades_filtered = 0
    consecutive_losses = 0
    max_consec_losses = 0
    current_consec_losses = 0

    trade_log = []
    equity_curve = []

    for i in range(200, len(df) - 20):
        row = df.iloc[i]
        current_time = df.index[i]
        in_test = current_time >= pd.Timestamp(test_start) and current_time < pd.Timestamp(test_end)

        month_str = current_time.strftime('%Y-%m')
        weekday = row['weekday']
        hour = row['hour']

        if weekday >= 5:
            continue

        # Track equity
        if in_test:
            current_equity = balance
            if position:
                if position['dir'] == 1:
                    unrealized = (row['close'] - position['entry']) * position['size']
                else:
                    unrealized = (position['entry'] - row['close']) * position['size']
                current_equity += unrealized
            equity_curve.append({'time': current_time, 'equity': current_equity, 'balance': balance})

        # Position management
        if position:
            exit_reason = None
            exit_price = None
            pnl = 0

            if (i - position['entry_idx']) >= MAX_HOLDING_HOURS:
                exit_price = row['close']
                pnl = (row['close'] - position['entry']) * position['size'] if position['dir'] == 1 else (position['entry'] - row['close']) * position['size']
                exit_reason = 'TIMEOUT'
            else:
                if position['dir'] == 1:
                    if row['low'] <= position['sl']:
                        exit_price = position['sl']
                        pnl = (position['sl'] - position['entry']) * position['size']
                        exit_reason = 'SL'
                    elif row['high'] >= position['tp']:
                        exit_price = position['tp']
                        pnl = (position['tp'] - position['entry']) * position['size']
                        exit_reason = 'TP'
                else:
                    if row['high'] >= position['sl']:
                        exit_price = position['sl']
                        pnl = (position['entry'] - position['sl']) * position['size']
                        exit_reason = 'SL'
                    elif row['low'] <= position['tp']:
                        exit_price = position['tp']
                        pnl = (position['entry'] - position['tp']) * position['size']
                        exit_reason = 'TP'

            if exit_reason:
                balance += pnl

                # Calculate pips
                if position['dir'] == 1:
                    pips = (exit_price - position['entry']) / 0.0001
                else:
                    pips = (position['entry'] - exit_price) / 0.0001

                if pnl > 0:
                    wins += 1
                    current_consec_losses = 0
                else:
                    losses += 1
                    current_consec_losses += 1
                    if current_consec_losses > max_consec_losses:
                        max_consec_losses = current_consec_losses

                if month_str not in monthly_pnl:
                    monthly_pnl[month_str] = {'pnl': 0, 'trades': 0, 'wins': 0}
                monthly_pnl[month_str]['pnl'] += pnl
                monthly_pnl[month_str]['trades'] += 1
                if pnl > 0:
                    monthly_pnl[month_str]['wins'] += 1

                # Log trade
                trade_log.append({
                    'ticket': len(trade_log) + 1,
                    'entry_time': position['entry_time'],
                    'exit_time': current_time,
                    'type': 'buy' if position['dir'] == 1 else 'sell',
                    'lots': round(position['size'] / 100000, 2),
                    'entry_price': position['entry'],
                    'sl': position['sl'],
                    'tp': position['tp'],
                    'exit_price': exit_price,
                    'pnl': pnl,
                    'pips': pips,
                    'exit_reason': exit_reason,
                    'rsi': position['rsi'],
                    'atr_pct': position['atr_pct'],
                    'regime': position['regime'],
                    'balance': balance
                })

                position = None
                consecutive_losses = current_consec_losses

            if balance > peak:
                peak = balance
            dd = (peak - balance) / peak * 100
            dd_amount = peak - balance
            if dd > max_dd:
                max_dd = dd
            if dd_amount > max_dd_amount:
                max_dd_amount = dd_amount

        # Entry logic with filters
        if not position and in_test:
            # Trading hours filter
            if hour < 7 or hour >= 22 or hour == 12:
                continue

            # ATR percentile filter
            atr_pct = row['atr_pct']
            if atr_pct < MIN_ATR_PCT or atr_pct > MAX_ATR_PCT:
                continue

            skip_trade = False

            # Regime filter
            if USE_REGIME_FILTER:
                if row['regime'] not in ALLOWED_REGIMES:
                    skip_trade = True
                    trades_filtered += 1

            # Consecutive loss filter
            if not skip_trade and USE_CONSEC_LOSS_FILTER:
                if consecutive_losses >= CONSEC_LOSS_LIMIT:
                    skip_trade = True
                    trades_filtered += 1
                    consecutive_losses = 0

            if skip_trade:
                continue

            # RSI signal
            rsi = row['rsi']
            signal = 1 if rsi < RSI_OVERSOLD else (-1 if rsi > RSI_OVERBOUGHT else 0)

            if signal:
                entry = row['close']
                atr = row['atr'] if row['atr'] > 0 else entry * 0.002

                # Dynamic TP
                base_tp = TP_LOW if atr_pct < 40 else (TP_HIGH if atr_pct > 60 else TP_MED)
                tp_mult = base_tp + 0.35 if 12 <= hour < 16 else base_tp

                # Calculate levels
                sl = entry - atr * SL_MULT if signal == 1 else entry + atr * SL_MULT
                tp = entry + atr * tp_mult if signal == 1 else entry - atr * tp_mult

                # Position sizing
                risk = balance * RISK_PER_TRADE
                size = min(risk / abs(entry - sl), 100000) if abs(entry - sl) > 0 else 0

                position = {
                    'dir': signal,
                    'entry': entry,
                    'sl': sl,
                    'tp': tp,
                    'size': size,
                    'entry_idx': i,
                    'entry_time': current_time,
                    'regime': row['regime'],
                    'rsi': rsi,
                    'atr_pct': atr_pct
                }
                trades_taken += 1

    total_trades = wins + losses
    win_rate = wins / total_trades * 100 if total_trades else 0
    total_return = (balance - INITIAL_BALANCE) / INITIAL_BALANCE * 100

    # Calculate additional stats
    winning_trades = [t for t in trade_log if t['pnl'] > 0]
    losing_trades = [t for t in trade_log if t['pnl'] <= 0]

    gross_profit = sum(t['pnl'] for t in winning_trades)
    gross_loss = abs(sum(t['pnl'] for t in losing_trades))
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float('inf')

    avg_win = gross_profit / len(winning_trades) if winning_trades else 0
    avg_loss = gross_loss / len(losing_trades) if losing_trades else 0

    return {
        'balance': balance,
        'total_return': total_return,
        'max_drawdown': max_dd,
        'max_dd_amount': max_dd_amount,
        'total_trades': total_trades,
        'wins': wins,
        'losses': losses,
        'win_rate': win_rate,
        'profit_factor': profit_factor,
        'gross_profit': gross_profit,
        'gross_loss': gross_loss,
        'avg_win': avg_win,
        'avg_loss': avg_loss,
        'max_consec_losses': max_consec_losses,
        'monthly_pnl': monthly_pnl,
        'trades_filtered': trades_filtered,
        'trade_log': trade_log,
        'equity_curve': equity_curve
    }


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(result, filename):
    """Export backtest results to Excel in MT5 Strategy Tester format."""

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0.00'
    pct_format = '0.00%'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')

    profit_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    loss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # =========================================================================
    # Sheet 1: Summary (like EA Report header)
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary['A1'] = "STRATEGY TESTER REPORT"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')

    ws_summary['A2'] = "RSI v3.7 OPTIMIZED (Python Backtest)"
    ws_summary['A2'].font = title_font
    ws_summary.merge_cells('A2:D2')

    # Settings section
    row = 4
    ws_summary[f'A{row}'] = "Settings"
    ws_summary[f'A{row}'].font = header_font
    row += 1

    settings = [
        ("Expert", "RSI_v37_OPTIMIZED"),
        ("Symbol", SYMBOL),
        ("Period", "H1 (2024.10.01 - 2026.01.31)"),
        ("Initial Deposit", f"${INITIAL_BALANCE:,.2f}"),
        ("Strategy", "SIDEWAYS Regime + ConsecLoss3"),
    ]

    for label, value in settings:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        row += 1

    # Parameters section
    row += 1
    ws_summary[f'A{row}'] = "Strategy Parameters"
    ws_summary[f'A{row}'].font = header_font
    row += 1

    params = [
        ("RSI Period", RSI_PERIOD),
        ("RSI Oversold", RSI_OVERSOLD),
        ("RSI Overbought", RSI_OVERBOUGHT),
        ("ATR Period", ATR_PERIOD),
        ("SL Multiplier", SL_MULT),
        ("TP Low/Med/High", f"{TP_LOW}/{TP_MED}/{TP_HIGH}"),
        ("Max Holding Hours", MAX_HOLDING_HOURS),
        ("ATR Range", f"{MIN_ATR_PCT}%-{MAX_ATR_PCT}%"),
        ("Risk Per Trade", f"{RISK_PER_TRADE*100}%"),
    ]

    for label, value in params:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = str(value)
        row += 1

    # Results section
    row += 1
    ws_summary[f'A{row}'] = "Results"
    ws_summary[f'A{row}'].font = header_font
    row += 1

    results = [
        ("Total Net Profit", f"${result['balance'] - INITIAL_BALANCE:,.2f}"),
        ("Total Return", f"{result['total_return']:.2f}%"),
        ("Gross Profit", f"${result['gross_profit']:,.2f}"),
        ("Gross Loss", f"${result['gross_loss']:,.2f}"),
        ("Profit Factor", f"{result['profit_factor']:.2f}"),
        ("Max Drawdown", f"{result['max_drawdown']:.2f}%"),
        ("Max Drawdown ($)", f"${result['max_dd_amount']:,.2f}"),
        ("Total Trades", result['total_trades']),
        ("Win Rate", f"{result['win_rate']:.2f}%"),
        ("Winners", result['wins']),
        ("Losers", result['losses']),
        ("Avg Win", f"${result['avg_win']:,.2f}"),
        ("Avg Loss", f"${result['avg_loss']:,.2f}"),
        ("Max Consec Losses", result['max_consec_losses']),
        ("Trades Filtered", result['trades_filtered']),
        ("Final Balance", f"${result['balance']:,.2f}"),
    ]

    for label, value in results:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = str(value)
        row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30

    # =========================================================================
    # Sheet 2: Monthly Breakdown
    # =========================================================================
    ws_monthly = wb.create_sheet("Monthly")

    ws_monthly['A1'] = "MONTHLY BREAKDOWN"
    ws_monthly['A1'].font = title_font
    ws_monthly.merge_cells('A1:E1')

    # Headers
    headers = ['Month', 'Trades', 'Wins', 'Win Rate', 'P/L']
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data
    row = 4
    for month, data in sorted(result['monthly_pnl'].items()):
        wr = (data['wins'] / data['trades'] * 100) if data['trades'] > 0 else 0

        ws_monthly.cell(row=row, column=1, value=month).border = thin_border
        ws_monthly.cell(row=row, column=2, value=data['trades']).border = thin_border
        ws_monthly.cell(row=row, column=3, value=data['wins']).border = thin_border
        ws_monthly.cell(row=row, column=4, value=f"{wr:.1f}%").border = thin_border

        pnl_cell = ws_monthly.cell(row=row, column=5, value=data['pnl'])
        pnl_cell.number_format = money_format
        pnl_cell.border = thin_border

        # Color code P/L
        if data['pnl'] >= 0:
            pnl_cell.fill = profit_fill
        else:
            pnl_cell.fill = loss_fill

        row += 1

    # Totals
    total_trades = sum(d['trades'] for d in result['monthly_pnl'].values())
    total_wins = sum(d['wins'] for d in result['monthly_pnl'].values())
    total_pnl = sum(d['pnl'] for d in result['monthly_pnl'].values())
    total_wr = (total_wins / total_trades * 100) if total_trades > 0 else 0

    ws_monthly.cell(row=row, column=1, value="TOTAL").font = header_font
    ws_monthly.cell(row=row, column=2, value=total_trades).font = header_font
    ws_monthly.cell(row=row, column=3, value=total_wins).font = header_font
    ws_monthly.cell(row=row, column=4, value=f"{total_wr:.1f}%").font = header_font
    ws_monthly.cell(row=row, column=5, value=total_pnl).font = header_font
    ws_monthly.cell(row=row, column=5).number_format = money_format

    # Column widths
    ws_monthly.column_dimensions['A'].width = 12
    ws_monthly.column_dimensions['B'].width = 10
    ws_monthly.column_dimensions['C'].width = 10
    ws_monthly.column_dimensions['D'].width = 12
    ws_monthly.column_dimensions['E'].width = 15

    # =========================================================================
    # Sheet 3: Trade List
    # =========================================================================
    ws_trades = wb.create_sheet("Trades")

    ws_trades['A1'] = "TRADE LIST"
    ws_trades['A1'].font = title_font
    ws_trades.merge_cells('A1:N1')

    # Headers
    trade_headers = ['#', 'Entry Time', 'Exit Time', 'Type', 'Lots', 'Entry', 'SL', 'TP',
                     'Exit', 'Pips', 'P/L', 'Reason', 'RSI', 'ATR%']

    for col, header in enumerate(trade_headers, 1):
        cell = ws_trades.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Trade data
    row = 4
    for trade in result['trade_log']:
        ws_trades.cell(row=row, column=1, value=trade['ticket']).border = thin_border
        ws_trades.cell(row=row, column=2, value=trade['entry_time'].strftime('%Y.%m.%d %H:%M')).border = thin_border
        ws_trades.cell(row=row, column=3, value=trade['exit_time'].strftime('%Y.%m.%d %H:%M')).border = thin_border
        ws_trades.cell(row=row, column=4, value=trade['type']).border = thin_border
        ws_trades.cell(row=row, column=5, value=trade['lots']).border = thin_border
        ws_trades.cell(row=row, column=6, value=round(trade['entry_price'], 5)).border = thin_border
        ws_trades.cell(row=row, column=7, value=round(trade['sl'], 5)).border = thin_border
        ws_trades.cell(row=row, column=8, value=round(trade['tp'], 5)).border = thin_border
        ws_trades.cell(row=row, column=9, value=round(trade['exit_price'], 5)).border = thin_border
        ws_trades.cell(row=row, column=10, value=round(trade['pips'], 1)).border = thin_border

        pnl_cell = ws_trades.cell(row=row, column=11, value=trade['pnl'])
        pnl_cell.number_format = money_format
        pnl_cell.border = thin_border
        if trade['pnl'] >= 0:
            pnl_cell.fill = profit_fill
        else:
            pnl_cell.fill = loss_fill

        ws_trades.cell(row=row, column=12, value=trade['exit_reason']).border = thin_border
        ws_trades.cell(row=row, column=13, value=round(trade['rsi'], 1)).border = thin_border
        ws_trades.cell(row=row, column=14, value=round(trade['atr_pct'], 0)).border = thin_border

        row += 1

    # Column widths for trades
    widths = [5, 18, 18, 6, 6, 10, 10, 10, 10, 8, 12, 10, 6, 6]
    for i, width in enumerate(widths, 1):
        ws_trades.column_dimensions[chr(64 + i)].width = width

    # =========================================================================
    # Sheet 4: Equity Curve Data
    # =========================================================================
    ws_equity = wb.create_sheet("Equity")

    ws_equity['A1'] = "EQUITY CURVE"
    ws_equity['A1'].font = title_font

    equity_headers = ['Time', 'Equity', 'Balance']
    for col, header in enumerate(equity_headers, 1):
        cell = ws_equity.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    # Sample every 24 hours for readability
    row = 4
    for i, eq in enumerate(result['equity_curve']):
        if i % 24 == 0:  # Sample daily
            ws_equity.cell(row=row, column=1, value=eq['time'].strftime('%Y-%m-%d'))
            ws_equity.cell(row=row, column=2, value=round(eq['equity'], 2))
            ws_equity.cell(row=row, column=3, value=round(eq['balance'], 2))
            row += 1

    ws_equity.column_dimensions['A'].width = 12
    ws_equity.column_dimensions['B'].width = 12
    ws_equity.column_dimensions['C'].width = 12

    # Save
    wb.save(filename)
    print(f"\nReport saved: {filename}")
    return filename


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("RSI v3.7 OPTIMIZED - Export Backtest Report")
    print("=" * 60)

    if not connect_mt5():
        return

    try:
        print(f"\nFetching {SYMBOL} H1 data...")
        start_date = datetime(2024, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(2026, 1, 31, 23, 59, tzinfo=timezone.utc)
        df = get_h1_data(SYMBOL, start_date, end_date)

        if df is None or len(df) == 0:
            print("Failed to get data")
            return

        print(f"Loaded {len(df)} bars")

        print("Calculating indicators...")
        df = prepare_indicators(df)

        print("\nRunning backtest...")
        result = run_backtest(df.copy())

        # Print summary
        print("\n" + "=" * 60)
        print("RESULTS SUMMARY")
        print("=" * 60)
        print(f"Total Return:    {result['total_return']:+.1f}%")
        print(f"Profit Factor:   {result['profit_factor']:.2f}")
        print(f"Max Drawdown:    {result['max_drawdown']:.1f}%")
        print(f"Win Rate:        {result['win_rate']:.1f}%")
        print(f"Total Trades:    {result['total_trades']}")

        # Export to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"C:/Users/Administrator/Music/SURGE-WSI/backtest/h1_strategy/RSI_v37_Optimized_Report_{timestamp}.xlsx"
        export_to_excel(result, filename)

        print("\nDone!")

    finally:
        mt5.shutdown()


if __name__ == "__main__":
    main()
